"""Excel (.xlsx) loader → common Document model."""

from __future__ import annotations

import logging
from pathlib import Path

from backend.loaders.base import ChapterHeading, Document

logger = logging.getLogger(__name__)


def load_excel(filepath: str | Path) -> Document:
    """Parse an .xlsx file into the common Document model.

    Each sheet becomes a chapter; each row is serialized as a text line.
    """
    filepath = Path(filepath)

    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel loading: pip install openpyxl")

    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    chapters: list[ChapterHeading] = []
    all_rows: list[str] = []

    for idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        chapters.append(ChapterHeading(
            title=sheet_name,
            level=1,
            order=idx,
        ))
        all_rows.append(f"## {sheet_name}")

        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join(str(cell) for cell in row if cell is not None)
            if row_text.strip():
                all_rows.append(row_text)

    wb.close()
    full_text = "\n".join(all_rows)

    return Document(
        filename=filepath.name,
        file_type="xlsx",
        chapters=chapters,
        full_text=full_text,
        char_count=len(full_text),
    )
